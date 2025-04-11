from openpyxl import load_workbook

def extract_categories_from_excel(excel_path) -> dict:
    # Загружаем Excel файл
    workbook = load_workbook(excel_path, read_only = True, data_only = True)
    sheet = workbook.active  # Предполагаем, что мы работаем с активным листом

    sub_categories = []

    # Пропускаем первые 4 строки, начинаем с пятой
    sub_category_current = None
    for row in sheet.iter_rows(min_row=5, values_only=True):
        sub_category = {
            "number": row[0],  # Столбец A
            "name": row[1],  # Столбец B
            "okpd_2": row[2],  # Столбец C
            "code": row[3],  # Столбец D
            "measure": row[4],  # Столбец E
            "chars": [], # Столбец F 5,6,7,8,9,10
            "category_name": row[11],  # Столбец L
            "ktru": row[12] if row[12] else None,  # Столбец M
            "kkn_code": row[13],  # Столбец N
            "product_part": row[14],  # Столбец O
            "actualization_date": row[15],  # Столбец P
            "russian": True if row[16] == "Да" else None  # Столбец Q (или 12)
        }
        if row[1]:
            sub_categories.append(sub_category)
            sub_category_current = sub_category
        # Заполнение характеристик
        characteristic = {
        "name": row[5],
        "value_1": row[6].replace(',', '.') if row[6] else None, # не менее
        "value_2": row[7].replace(',', '.') if row[7] else None, # не более
        "value_3": None, # варианты значений
        "value_4": row[9] if row[9] else None, # одно значение
        "measure": row[10] if row[10] else None,
        }
        if row[8]:
            characteristic['value_3'] = [v for v in row[8].split(';')]
        if row[5]:
            sub_category_current["chars"].append(characteristic)

    return sub_categories

# import datetime
# sub_categories_example = [{
#     'id': 1, 'number': 2486, 'name': 'Кресло офисное¹ тип 15',
#     'okpd_2': '31.01.12.160', 'code': '028', 'measure': 'ШТ',
#     'chars': [{'name': 'Назначение', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Для руководителя', 'measure': '-'}, {'name': 'Тип каркаса', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Деревянный', 'measure': '-'}, {'name': 'Вид материала обивки сидения', 'value_1': None, 'value_2': None, 'value_3': ['Экокожа', ' Текстиль', ' Кожа искусственная'], 'value_4': None, 'measure': '-'}, {'name': 'Вид материала обивки спинки', 'value_1': None, 'value_2': None, 'value_3': ['Экокожа', ' Текстиль', ' Кожа искусственная'], 'value_4': None, 'measure': '-'}, {'name': 'Наличие механизма регулировки по высоте', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Нет', 'measure': '-'}, {'name': 'Регулировка угла наклона', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Нет', 'measure': '-'}, {'name': 'Наличие подлокотников', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Да', 'measure': '-'}, {'name': 'Ширина', 'value_1': '840', 'value_2': '1020', 'value_3': None, 'value_4': None, 'measure': 'ММ'}, {'name': 'Глубина', 'value_1': '770', 'value_2': '950', 'value_3': None, 'value_4': None, 'measure': 'ММ'}, {'name': 'Высота', 'value_1': '730', 'value_2': '900', 'value_3': None, 'value_4': None, 'measure': 'ММ'}, {'name': 'Высота сиденья', 'value_1': '370', 'value_2': '450', 'value_3': None, 'value_4': None, 'measure': 'ММ'}, {'name': 'Глубина сиденья', 'value_1': '500', 'value_2': '560', 'value_3': None, 'value_4': None, 'measure': 'ММ'}, {'name': 'Тип напольной опоры', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Ножки', 'measure': '-'}, {'name': 'Соответствие нормативно-технической документации', 'value_1': None, 'value_2': None, 'value_3': ['ТР ТС 025/2012', ' ГОСТ 16371-2014', ' ГОСТ 19917-2014', ' ТУ производителя'], 'value_4': None, 'measure': '-'}],
#     'category_name': 'Кресло мягкое на деревянном каркасе',
#     'ktru': '31.01.12.160-00000005', 'kkn_code': '31.01.12.160-028',
#     'product_part': '05. Мебель',
#     'actualization_date': datetime.datetime(2024, 3, 29, 0, 0), 'russian': True},
#     {'id': 2, 'number': 4486, 'name': 'Упаковка для стерилизации, одноразового использования¹ тип 2',
#     'okpd_2': '32.50.50.190', 'code': '125', 'measure': 'ШТ',
#     'chars': [{'name': 'Вид стерилизации', 'value_1': None, 'value_2': None, 'value_3': ['Паровая, Газовая, Радиационная, Воздушная', ' Паровая, Газовая', ' Паровая', ' Газовая', ' Радиационная', ' Воздушная'], 'value_4': None, 'measure': '-'}, {'name': 'Тип', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'Бумага крепированная', 'measure': '-'}, {'name': 'Форма выпуска', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': 'В листах', 'measure': '-'}, {'name': 'Плотность бумажного слоя площадью 1м2, г', 'value_1': '58', 'value_2': None, 'value_3': None, 'value_4': None, 'measure': '-'}, {'name': 'Цвет бумаги', 'value_1': None, 'value_2': None, 'value_3': ['Белая', ' Цветная', ' Двухцветная'], 'value_4': None, 'measure': '-'}, {'name': 'Ширина листа', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': '1000', 'measure': 'ММ'}, {'name': 'Длина листа', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': '1000', 'measure': 'ММ'}, {'name': 'Номенклатурная классификация медицинских изделий по видам', 'value_1': None, 'value_2': None, 'value_3': None, 'value_4': '185910', 'measure': '-'}],
#     'category_name': 'Бумага одноразового использования для стерилизации',
#     'ktru': '32.50.50.190-00000337', 'kkn_code': '32.50.50.190-125',
#     'product_part': '12. Средства для дезинфекции и стерилизации',
#     'actualization_date': datetime.datetime(2021, 9, 30, 0, 0), 'russian': True}
# ]

# for k,v in sub_categories_example.items():
#     if type(v) != list:
#         print([k, v])
#     else:
#         for chars in v:
#             print(' - - ',chars)
# print()
